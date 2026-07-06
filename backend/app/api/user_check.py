"""Bulk user-existence check: upload CSV/XLSX, split rows into exists/missing buckets."""
import asyncio
import csv
import io
import logging
import time
from datetime import date, datetime
from typing import Any, Literal

import httpx
from fastapi import APIRouter, Form, HTTPException, Request, UploadFile
from fastapi.responses import JSONResponse

from app.lightspeed_api import list_users

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/user-check", tags=["user-check"])

_RATE_LIMIT = (10, 60)
_rate_timestamps: dict[str, list[float]] = {}

_MAX_ROWS = 20_000
_MAX_FILE_BYTES = 20 * 1024 * 1024  # 20 MB safety cap

_CSV_MIMES = {"text/csv", "application/csv", "text/plain", ""}
_XLSX_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/octet-stream",
}


def _has_credentials(session: dict) -> bool:
    return bool(session.get("api_key") and session.get("api_secret"))


def _check_rate_limit(client_ip: str) -> bool:
    max_req, window = _RATE_LIMIT
    now = time.monotonic()
    if client_ip not in _rate_timestamps:
        _rate_timestamps[client_ip] = []
    timestamps = _rate_timestamps[client_ip]
    timestamps[:] = [t for t in timestamps if now - t < window]
    if len(timestamps) >= max_req:
        return False
    timestamps.append(now)
    return True


def _detect_format(file: UploadFile) -> Literal["csv", "xlsx"]:
    mime = (file.content_type or "").lower()
    name = (file.filename or "").lower()
    if mime in _XLSX_MIMES or name.endswith(".xlsx"):
        return "xlsx"
    if mime in _CSV_MIMES or name.endswith(".csv"):
        return "csv"
    raise HTTPException(
        status_code=415,
        detail=f"Unsupported file type (content-type={mime!r}, filename={file.filename!r}). Upload a .csv or .xlsx file.",
    )


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def _parse_csv(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    # Strip BOM if present, then decode as UTF-8 with a permissive fallback.
    text: str
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    reader = csv.reader(io.StringIO(text))
    headers: list[str] = []
    rows: list[dict[str, str]] = []
    for i, cells in enumerate(reader):
        if i == 0:
            headers = [str(c).strip() for c in cells]
            continue
        if not cells:
            continue
        row: dict[str, str] = {}
        for j, h in enumerate(headers):
            row[h] = cells[j].strip() if j < len(cells) else ""
        rows.append(row)
        if len(rows) > _MAX_ROWS:
            raise HTTPException(status_code=413, detail=f"CSV has more than {_MAX_ROWS} rows")
    return headers, rows


def _parse_xlsx(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    from openpyxl import load_workbook  # local import so the module still loads without the dep during dev

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read XLSX: {e!s}") from e
    try:
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="XLSX has no active sheet")
        headers: list[str] = []
        rows: list[dict[str, str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [_cell_to_str(c).strip() for c in row]
                continue
            if row is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                continue
            record: dict[str, str] = {}
            for j, h in enumerate(headers):
                record[h] = _cell_to_str(row[j]).strip() if j < len(row) else ""
            rows.append(record)
            if len(rows) > _MAX_ROWS:
                raise HTTPException(status_code=413, detail=f"XLSX has more than {_MAX_ROWS} rows")
        return headers, rows
    finally:
        wb.close()


async def _parse_upload(file: UploadFile) -> tuple[Literal["csv", "xlsx"], list[str], list[dict[str, str]]]:
    fmt = _detect_format(file)
    data = await file.read()
    if len(data) > _MAX_FILE_BYTES:
        raise HTTPException(status_code=413, detail="File is larger than the 20 MB limit")
    if not data:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if fmt == "csv":
        headers, rows = _parse_csv(data)
    else:
        headers, rows = _parse_xlsx(data)
    headers = [h for h in headers if h]  # drop empty header cells at the end
    if not headers:
        raise HTTPException(status_code=400, detail="No header row detected in the uploaded file")
    return fmt, headers, rows


def _norm_email(v: Any) -> str:
    return str(v or "").strip().lower()


def _norm_username(v: Any) -> str:
    return str(v or "").strip().lower()


def _norm_user_id(v: Any) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    try:
        # Coerce "12345", "12345.0", 12345, 12345.0 all to "12345"
        return str(int(float(s)))
    except (TypeError, ValueError):
        return s


def _build_index(
    users: list[dict[str, Any]], key_type: Literal["email", "username", "userId"]
) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for u in users:
        if key_type == "email":
            key = _norm_email(u.get("email"))
        elif key_type == "username":
            key = _norm_username(u.get("username"))
        else:
            key = _norm_user_id(u.get("id"))
        if key and key not in index:
            index[key] = u
    return index


def _normalize_key(value: Any, key_type: Literal["email", "username", "userId"]) -> str:
    if key_type == "email":
        return _norm_email(value)
    if key_type == "username":
        return _norm_username(value)
    return _norm_user_id(value)


@router.post("/preview")
async def preview_upload(request: Request, file: UploadFile):
    """
    Parse just the headers of a CSV/XLSX upload. Requires session credentials.
    Returns { format, headers, row_count }.
    """
    client_ip = request.client.host if request.client else "unknown"
    if not _check_rate_limit(client_ip):
        return JSONResponse(status_code=429, content={"detail": "Too many requests"})
    session = request.state.session
    if not _has_credentials(session):
        raise HTTPException(status_code=401, detail="No credentials in session")
    fmt, headers, rows = await _parse_upload(file)
    return {"format": fmt, "headers": headers, "row_count": len(rows)}


@router.post("/check")
async def check_users(
    request: Request,
    file: UploadFile,
    lookup_column: str = Form(..., min_length=1),
    lookup_key_type: Literal["email", "username", "userId"] = Form(...),
):
    """
    Split the uploaded rows into 'exists' and 'missing' buckets using an index built from LSVT's user list.
    """
    client_ip = request.client.host if request.client else "unknown"
    if not _check_rate_limit(client_ip):
        return JSONResponse(status_code=429, content={"detail": "Too many requests"})
    session = request.state.session
    if not _has_credentials(session):
        raise HTTPException(status_code=401, detail="No credentials in session")

    fmt, headers, rows = await _parse_upload(file)
    if lookup_column not in headers:
        raise HTTPException(
            status_code=400,
            detail=f"lookup_column {lookup_column!r} not found in file headers",
        )

    api_key = session["api_key"]
    api_secret = session["api_secret"]
    all_users: list[dict[str, Any]] = []
    page = 1
    try:
        while True:
            batch = await list_users(api_key, api_secret, items_per_page=200, page=page)
            if not batch:
                break
            all_users.extend(batch)
            if len(batch) < 200:
                break
            page += 1
            await asyncio.sleep(0.2)
    except httpx.HTTPStatusError as e:
        status = e.response.status_code if e.response is not None else 502
        raise HTTPException(
            status_code=502 if status >= 500 else status,
            detail=f"LightSpeed VT API error while listing users: {e!s}",
        ) from e
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"LightSpeed VT API error: {e!s}") from e

    index = _build_index(all_users, lookup_key_type)
    logger.info(
        "user-check load lsvt_users=%d indexed=%d key_type=%s",
        len(all_users), len(index), lookup_key_type,
    )

    exists: list[dict[str, Any]] = []
    missing: list[dict[str, Any]] = []
    blank_lookup = 0
    for row in rows:
        raw = row.get(lookup_column, "")
        key = _normalize_key(raw, lookup_key_type)
        if not key:
            blank_lookup += 1
            missing.append({**row, "reason": "lookup value empty"})
            continue
        match = index.get(key)
        if match is None:
            missing.append({**row, "reason": "not found in system"})
        else:
            exists.append({
                **row,
                "matched_userId": str(match.get("id", "") or ""),
                "matched_username": str(match.get("username", "") or ""),
                "matched_email": str(match.get("email", "") or ""),
            })

    return {
        "lookup_key_type": lookup_key_type,
        "lookup_column": lookup_column,
        "format": fmt,
        "headers": headers,
        "extra_headers_exists": ["matched_userId", "matched_username", "matched_email"],
        "extra_headers_missing": ["reason"],
        "exists": exists,
        "missing": missing,
        "summary": {
            "total": len(rows),
            "exists": len(exists),
            "missing": len(missing),
            "blank_lookup": blank_lookup,
        },
    }
